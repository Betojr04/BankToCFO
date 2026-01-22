from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from services.categorizer import get_category_summary, get_monthly_summary
from datetime import datetime


def generate_cfo_pack(transactions: List[Dict], output_path: Path):
    """
    Generate State-of-the-Art CFO Dashboard
    """

    # Calculate metrics
    total_income = sum(t["amount"] for t in transactions if t["amount"] > 0)
    total_expenses = abs(sum(t["amount"] for t in transactions if t["amount"] < 0))
    net_income = total_income - total_expenses

    dates = [datetime.strptime(t["date"], "%Y-%m-%d") for t in transactions]
    start_date = min(dates)
    end_date = max(dates)
    months_count = len(set(d.strftime("%Y-%m") for d in dates))

    monthly_summary = get_monthly_summary(transactions)
    category_summary = get_category_summary(transactions)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets
    create_executive_dashboard(
        wb,
        transactions,
        total_income,
        total_expenses,
        net_income,
        start_date,
        end_date,
        months_count,
        monthly_summary,
        category_summary,
    )

    create_transactions_sheet(wb, transactions)
    create_monthly_sheet(wb, monthly_summary)
    create_category_sheet(wb, category_summary, total_expenses)
    create_instructions_sheet(wb)

    # Save workbook
    wb.save(output_path)


def create_executive_dashboard(
    wb,
    transactions,
    total_income,
    total_expenses,
    net_income,
    start_date,
    end_date,
    months_count,
    monthly_summary,
    category_summary,
):
    """Create professional executive dashboard"""

    ws = wb.create_sheet("Executive Dashboard", 0)

    # Set column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 3
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 15

    # ===== HEADER SECTION =====
    ws.merge_cells("B2:D2")
    ws["B2"] = "Financial Dashboard"
    ws["B2"].font = Font(size=24, bold=True, color="1E3A8A")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B3:D3")
    ws["B3"] = f'{start_date.strftime("%b %d, %Y")} - {end_date.strftime("%b %d, %Y")}'
    ws["B3"].font = Font(size=11, color="64748B")
    ws["B3"].alignment = Alignment(horizontal="left")

    # ===== KPI CARDS =====
    create_kpi_card(ws, "B5", "Total Income", total_income, "059669", is_currency=True)
    create_kpi_card(
        ws, "C5", "Total Expenses", total_expenses, "DC2626", is_currency=True
    )

    # Net Income card (green if positive, red if negative)
    color = "059669" if net_income >= 0 else "DC2626"
    create_kpi_card(ws, "D5", "Net Income", net_income, color, is_currency=True)

    # Additional metrics
    ws["B8"] = "Period"
    ws["B8"].font = Font(size=10, color="64748B")
    ws["C8"] = f"{months_count} months"
    ws["C8"].font = Font(size=10, bold=True)
    ws["C8"].alignment = Alignment(horizontal="left")

    ws["B9"] = "Transactions"
    ws["B9"].font = Font(size=10, color="64748B")
    ws["C9"] = f"{len(transactions)}"
    ws["C9"].font = Font(size=10, bold=True)
    ws["C9"].alignment = Alignment(horizontal="left")

    # ===== RECENT TRANSACTIONS SECTION =====
    ws["B12"] = "Recent Transactions"
    ws["B12"].font = Font(size=14, bold=True, color="1E3A8A")

    # Headers
    headers = ["Date", "Description", "Category", "Amount"]
    header_fill = PatternFill(
        start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"
    )

    for idx, header in enumerate(headers):
        cell = ws.cell(row=14, column=idx + 2)
        cell.value = header
        cell.font = Font(size=10, bold=True, color="1E3A8A")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if idx < 2 else "right")

    # Recent transactions (last 7)
    recent = sorted(transactions, key=lambda x: x["date"], reverse=True)[:7]
    for idx, trans in enumerate(recent):
        row = 15 + idx
        ws.cell(row=row, column=2).value = trans["date"]
        ws.cell(row=row, column=3).value = trans["description"][:40]
        ws.cell(row=row, column=4).value = trans["category"]
        ws.cell(row=row, column=5).value = trans["amount"]
        ws.cell(row=row, column=5).number_format = "$#,##0.00"

        # Color code amounts
        amount_cell = ws.cell(row=row, column=5)
        if trans["amount"] < 0:
            amount_cell.font = Font(color="DC2626", bold=True)
        else:
            amount_cell.font = Font(color="059669", bold=True)

        # Right align amounts
        for col in [4, 5]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")

    # ===== CATEGORY BREAKDOWN SECTION =====
    ws["F12"] = "By Category"
    ws["F12"].font = Font(size=14, bold=True, color="1E3A8A")

    # Category headers
    ws["F14"] = "Category"
    ws["G14"] = "Amount"
    for cell in [ws["F14"], ws["G14"]]:
        cell.font = Font(size=10, bold=True, color="1E3A8A")
        cell.fill = header_fill

    ws["G14"].alignment = Alignment(horizontal="right")

    # Top 7 expense categories
    expenses = {k: abs(v) for k, v in category_summary.items() if v < 0}
    top_expenses = sorted(expenses.items(), key=lambda x: x[1], reverse=True)[:7]

    for idx, (cat, amt) in enumerate(top_expenses):
        row = 15 + idx
        ws.cell(row=row, column=6).value = cat
        ws.cell(row=row, column=7).value = amt
        ws.cell(row=row, column=7).number_format = "$#,##0.00"
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=7).font = Font(color="DC2626", bold=True)

    # Add data bars to category amounts
    data_bar_rule = DataBarRule(
        start_type="min", end_type="max", color="DC2626", showValue=True
    )
    ws.conditional_formatting.add(f"G15:G{14+len(top_expenses)}", data_bar_rule)

    # ===== MONTHLY CASHFLOW SECTION =====
    ws["B25"] = "Monthly Cashflow"
    ws["B25"].font = Font(size=14, bold=True, color="1E3A8A")

    # Create monthly data table for chart
    chart_row_start = 27
    ws.cell(chart_row_start, 2).value = "Month"
    ws.cell(chart_row_start, 3).value = "Income"
    ws.cell(chart_row_start, 4).value = "Expenses"

    for cell in [
        ws.cell(chart_row_start, 2),
        ws.cell(chart_row_start, 3),
        ws.cell(chart_row_start, 4),
    ]:
        cell.font = Font(size=10, bold=True, color="1E3A8A")
        cell.fill = header_fill

    # Add monthly data
    data_row = chart_row_start + 1
    for month, data in sorted(monthly_summary.items()):
        ws.cell(data_row, 2).value = month
        ws.cell(data_row, 3).value = data["revenue"]
        ws.cell(data_row, 4).value = data["expenses"]
        ws.cell(data_row, 3).number_format = "$#,##0"
        ws.cell(data_row, 4).number_format = "$#,##0"
        data_row += 1

    # Create bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "Income vs Expenses"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = None
    chart.height = 10
    chart.width = 16
    chart.legend = None

    data = Reference(
        ws, min_col=3, min_row=chart_row_start, max_row=data_row - 1, max_col=4
    )
    cats = Reference(ws, min_col=2, min_row=chart_row_start + 1, max_row=data_row - 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Style chart
    from openpyxl.chart.series import DataLabelList

    for i, series in enumerate(chart.series):
        series.graphicalProperties.solidFill = "059669" if i == 0 else "DC2626"

    ws.add_chart(chart, f"B{data_row + 2}")


def create_kpi_card(ws, cell, label, value, color, is_currency=False):
    """Create a visual KPI card"""

    row = int(cell[1:])
    col_letter = cell[0]

    # Card background
    fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # Label
    label_cell = ws[f"{col_letter}{row}"]
    label_cell.value = label
    label_cell.font = Font(size=9, color="64748B")
    label_cell.fill = fill
    label_cell.border = border
    label_cell.alignment = Alignment(horizontal="left", vertical="top")

    # Value
    value_cell = ws[f"{col_letter}{row+1}"]
    if is_currency:
        value_cell.value = value
        value_cell.number_format = "$#,##0"
    else:
        value_cell.value = value

    value_cell.font = Font(size=18, bold=True, color=color)
    value_cell.fill = fill
    value_cell.border = border
    value_cell.alignment = Alignment(horizontal="left", vertical="bottom")


def create_transactions_sheet(wb, transactions):
    """Create detailed transactions sheet"""

    ws = wb.create_sheet("Transactions")

    # Headers
    headers = ["Date", "Description", "Category", "Amount", "Type"]
    header_fill = PatternFill(
        start_color="0D9488", end_color="0D9488", fill_type="solid"
    )

    for idx, header in enumerate(headers, 1):
        cell = ws.cell(1, idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 10

    # Data
    df = pd.DataFrame(transactions)
    df = df[["date", "description", "category", "amount", "type"]]
    df = df.sort_values("date", ascending=False)

    for idx, row in enumerate(df.itertuples(index=False), 2):
        ws.cell(idx, 1).value = row.date
        ws.cell(idx, 2).value = row.description
        ws.cell(idx, 3).value = row.category
        ws.cell(idx, 4).value = row.amount
        ws.cell(idx, 4).number_format = "$#,##0.00"
        ws.cell(idx, 5).value = row.type

        # Color code amounts
        if row.amount < 0:
            ws.cell(idx, 4).font = Font(color="DC2626", bold=True)
        else:
            ws.cell(idx, 4).font = Font(color="059669", bold=True)

    ws.freeze_panes = "A2"


def create_monthly_sheet(wb, monthly_summary):
    """Create monthly analysis sheet"""

    ws = wb.create_sheet("Monthly Analysis")

    headers = ["Month", "Income", "Expenses", "Net Income"]
    header_fill = PatternFill(
        start_color="0D9488", end_color="0D9488", fill_type="solid"
    )

    for idx, header in enumerate(headers, 1):
        cell = ws.cell(1, idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 15

    for idx, (month, data) in enumerate(sorted(monthly_summary.items()), 2):
        ws.cell(idx, 1).value = month
        ws.cell(idx, 2).value = data["revenue"]
        ws.cell(idx, 3).value = data["expenses"]
        ws.cell(idx, 4).value = data["net_income"]

        for col in [2, 3, 4]:
            ws.cell(idx, col).number_format = "$#,##0.00"

        # Color code net income
        if data["net_income"] < 0:
            ws.cell(idx, 4).font = Font(color="DC2626", bold=True)
        else:
            ws.cell(idx, 4).font = Font(color="059669", bold=True)


def create_category_sheet(wb, category_summary, total_expenses):
    """Create category analysis sheet"""

    ws = wb.create_sheet("Category Analysis")

    headers = ["Category", "Amount", "Percentage"]
    header_fill = PatternFill(
        start_color="0D9488", end_color="0D9488", fill_type="solid"
    )

    for idx, header in enumerate(headers, 1):
        cell = ws.cell(1, idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15

    categories = sorted(category_summary.items(), key=lambda x: abs(x[1]), reverse=True)

    for idx, (cat, amt) in enumerate(categories, 2):
        ws.cell(idx, 1).value = cat
        ws.cell(idx, 2).value = amt
        ws.cell(idx, 2).number_format = "$#,##0.00"

        if amt < 0:
            pct = (abs(amt) / total_expenses * 100) if total_expenses > 0 else 0
            ws.cell(idx, 3).value = pct / 100
            ws.cell(idx, 3).number_format = "0.0%"
            ws.cell(idx, 2).font = Font(color="DC2626")
        else:
            ws.cell(idx, 3).value = 0
            ws.cell(idx, 3).number_format = "0.0%"
            ws.cell(idx, 2).font = Font(color="059669")


def create_instructions_sheet(wb):
    """Create how to use sheet"""

    ws = wb.create_sheet("How to Use")

    ws.column_dimensions["A"].width = 80

    content = [
        ("BankToCFO - Professional Financial Dashboard", 20, True, "0D9488"),
        ("", 10, False, None),
        ("What You're Looking At:", 14, True, "1E3A8A"),
        (
            "This CFO Pack transforms your bank statement into an executive-level financial dashboard with AI-powered categorization and professional visualizations.",
            10,
            False,
            None,
        ),
        ("", 10, False, None),
        ("Sheet Breakdown:", 14, True, "1E3A8A"),
        ("", 10, False, None),
        (
            "1. Executive Dashboard - Your command center with KPIs, recent transactions, and category breakdown",
            10,
            False,
            None,
        ),
        (
            "2. Transactions - Complete list of all transactions with AI categorization",
            10,
            False,
            None,
        ),
        (
            "3. Monthly Analysis - Month-over-month income, expenses, and net income trends",
            10,
            False,
            None,
        ),
        (
            "4. Category Analysis - Detailed spending breakdown by category with percentages",
            10,
            False,
            None,
        ),
        ("5. How to Use - This guide", 10, False, None),
        ("", 10, False, None),
        ("Key Metrics Explained:", 14, True, "1E3A8A"),
        ("", 10, False, None),
        (
            "Total Income - All money coming in (salary, deposits, payments received)",
            10,
            False,
            None,
        ),
        (
            "Total Expenses - All money going out (purchases, bills, transfers)",
            10,
            False,
            None,
        ),
        ("Net Income - Income minus Expenses (your bottom line)", 10, False, None),
        ("", 10, False, None),
        ("Category Intelligence:", 14, True, "1E3A8A"),
        ("", 10, False, None),
        (
            "Our AI categorizes transactions into 30+ categories including Fast Food, Restaurants, Gas & Fuel, Subscriptions, Shopping, Debt Payments, Fitness, Healthcare, and more. Categorization is 85-90% accurate.",
            10,
            False,
            None,
        ),
        ("", 10, False, None),
        ("Pro Tips:", 14, True, "1E3A8A"),
        ("", 10, False, None),
        (
            "• Review the Executive Dashboard first for your financial snapshot",
            10,
            False,
            None,
        ),
        ('• Check "By Category" to identify spending opportunities', 10, False, None),
        ("• Use Monthly Analysis to spot trends over time", 10, False, None),
        ("• Verify transaction categories for tax purposes", 10, False, None),
        ("• Upload multiple months to track long-term patterns", 10, False, None),
        ("", 10, False, None),
        ("Questions? support@banktocfo.com", 11, True, "0D9488"),
    ]

    for idx, item in enumerate(content, 1):
        text, size, bold, color = item
        cell = ws.cell(idx, 1)
        cell.value = text
        cell.font = Font(size=size, bold=bold, color=color if color else "000000")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        if size > 12:
            ws.row_dimensions[idx].height = 25
